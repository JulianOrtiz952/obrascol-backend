import openpyxl
from io import BytesIO
from .models import Bodega, Subbodega, Material, Marca, Factura, Movimiento, UnidadMedida

def export_all_data_to_excel(template=False):
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # helper for adding sheets
    def add_sheet(name, queryset, fields, header_names=None):
        ws = wb.create_sheet(title=name)
        headers = header_names if header_names else [f.replace('_', ' ').title() for f in fields]
        ws.append(headers)
        
        if not template:
            for obj in queryset:
                row = []
                for field in fields:
                    val = getattr(obj, field)
                    # Handle foreign keys or special types
                    if hasattr(val, '__str__') and not isinstance(val, (str, int, float, bool, type(None))):
                        row.append(str(val))
                    else:
                        row.append(val)
                ws.append(row)

    # 1. Bodegas
    add_sheet(
        "Bodegas", 
        Bodega.objects.all(), 
        ['id', 'nombre', 'ubicacion', 'activo']
    )

    # 2. Subbodegas
    add_sheet(
        "Subbodegas",
        Subbodega.objects.select_related('bodega', 'parent').all(),
        ['id', 'nombre', 'bodega', 'parent', 'activo'],
        ['ID', 'Nombre', 'Bodega Padre', 'Subbodega Padre', 'Activo']
    )

    # 3. Materiales
    add_sheet(
        "Materiales",
        Material.objects.select_related('marca').all(),
        ['id', 'codigo', 'codigo_barras', 'referencia', 'nombre', 'unidad', 'marca'],
        ['ID', 'Código', 'Código Barras', 'Referencia', 'Nombre', 'Unidad', 'Marca']
    )

    # 4. Marcas
    add_sheet(
        "Marcas",
        Marca.objects.all(),
        ['id', 'nombre', 'activo']
    )

    # 5. Facturas
    add_sheet(
        "Facturas",
        Factura.objects.all(),
        ['id', 'numero', 'proveedor', 'fecha']
    )

    # 6. Movimientos (Kardex - All)
    add_sheet(
        "Movimientos",
        Movimiento.objects.select_related(
            'material', 'bodega', 'subbodega', 'bodega_destino', 
            'subbodega_destino', 'marca', 'factura', 'usuario'
        ).all().order_by('-fecha'),
        ['id', 'fecha', 'tipo', 'material', 'cantidad', 'bodega', 'subbodega', 'bodega_destino', 'subbodega_destino', 'marca', 'factura_manual', 'observaciones', 'usuario'],
        ['ID', 'Fecha', 'Tipo', 'Material', 'Cantidad', 'Bodega', 'Subbodega', 'Bodega Destino', 'Subbodega Destino', 'Marca', 'Factura Manual', 'Observaciones', 'Usuario']
    )

    # 7. Specialized sheets for Template
    move_fields = ['id', 'fecha', 'material', 'cantidad', 'bodega', 'subbodega', 'marca', 'factura_manual', 'observaciones']
    move_headers = ['ID', 'Fecha', 'Material', 'Cantidad', 'Bodega', 'Subbodega', 'Marca', 'Factura Manual', 'Observaciones']
    
    traslado_fields = ['id', 'fecha', 'material', 'cantidad', 'bodega', 'subbodega', 'bodega_destino', 'subbodega_destino', 'marca', 'observaciones']
    traslado_headers = ['ID', 'Fecha', 'Material', 'Cantidad', 'Bodega Origen', 'Subbodega Origen', 'Bodega Destino', 'Subbodega Destino', 'Marca', 'Observaciones']

    def add_specialized_sheet(name, tipo):
        ws = wb.create_sheet(title=name)
        headers = traslado_headers if tipo == 'Traslado' else move_headers
        fields = traslado_fields if tipo == 'Traslado' else move_fields
        ws.append(headers)
        
        if template:
            # Add an example row
            example_row = []
            for h in headers:
                if h == 'ID': example_row.append("EJEMPLO (BORRAR)")
                elif h == 'Fecha': example_row.append(timezone.now().strftime('%Y-%m-%d %H:%M'))
                elif h == 'Material': example_row.append("CODIGO - NOMBRE")
                elif h == 'Cantidad': example_row.append(1)
                elif h == 'Marca': example_row.append("Generico")
                elif h.startswith('Bodega'): example_row.append("PUNTO DE VENTA")
                elif h.startswith('Subbodega'): example_row.append("GENERAL")
                else: example_row.append("")
            ws.append(example_row)
        else:
            # Query existing movements of this type
            queryset = Movimiento.objects.filter(tipo=tipo).select_related(
                'material', 'bodega', 'subbodega', 'bodega_destino', 
                'subbodega_destino', 'marca'
            ).order_by('-fecha')
            for obj in queryset:
                row = []
                for field in fields:
                    if field == 'bodega' and tipo == 'Traslado': val = obj.bodega
                    elif field == 'subbodega' and tipo == 'Traslado': val = obj.subbodega
                    else: val = getattr(obj, field)
                    
                    if hasattr(val, '__str__') and not isinstance(val, (str, int, float, bool, type(None))):
                        row.append(str(val))
                    else:
                        row.append(val)
                ws.append(row)

    add_specialized_sheet("Entradas", "Entrada")
    add_specialized_sheet("Salidas", "Salida")
    add_specialized_sheet("Traslados", "Traslado")

    wb.save(output)
    output.seek(0)
    return output
from django.db import transaction
from django.utils import timezone
import datetime

def import_all_data_from_excel(file_ptr, user=None):
    wb = openpyxl.load_workbook(file_ptr)
    summary = {"created": 0, "updated": 0, "errors": []}

    def get_clean_val(row, headers, col_name):
        # Flexible matching: case-insensitive and stripped
        col_name_norm = col_name.lower().strip()
        headers_norm = [str(h).lower().strip() for h in headers]
        try:
            # Try exact match first
            if col_name_norm in headers_norm:
                idx = headers_norm.index(col_name_norm)
            else:
                # Try partial match if exactly one header starts with the name (handles "Subbodeg")
                matches = [i for i, h in enumerate(headers_norm) if h.startswith(col_name_norm)]
                if len(matches) == 1:
                    idx = matches[0]
                else:
                    return None
            
            val = row[idx]
            if val == 'None' or val == '':
                return None
            return val
        except (ValueError, IndexError):
            return None

    # We import in a specific order to handle dependencies
    # 1. Marcas
    # 2. Bodegas
    # 3. Subbodegas
    # 4. Materiales
    # 5. Facturas
    # 6. Movimientos

    with transaction.atomic():
        # --- Marcas ---
        if "Marcas" in wb.sheetnames:
            ws = wb["Marcas"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "id")
                    nombre = get_clean_val(row, headers, "nombre")
                    activo = get_clean_val(row, headers, "activo")
                    
                    if nombre:
                        defaults = {'nombre': nombre, 'activo': bool(activo) if activo is not None else True}
                        if id_val:
                            obj, created = Marca.objects.update_or_create(id=id_val, defaults=defaults)
                        else:
                            # Natural key fallback: nombre
                            obj, created = Marca.objects.update_or_create(nombre=nombre, defaults=defaults)
                        
                        if created: summary["created"] += 1
                        else: summary["updated"] += 1

        # --- Bodegas ---
        if "Bodegas" in wb.sheetnames:
            ws = wb["Bodegas"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "id")
                    nombre = get_clean_val(row, headers, "nombre")
                    ubicacion = get_clean_val(row, headers, "ubicacion") or ""
                    activo = get_clean_val(row, headers, "activo")
                    
                    if nombre:
                        defaults = {'nombre': nombre, 'ubicacion': ubicacion, 'activo': bool(activo) if activo is not None else True}
                        if id_val:
                            obj, created = Bodega.objects.update_or_create(id=id_val, defaults=defaults)
                        else:
                            # Natural key fallback: nombre
                            obj, created = Bodega.objects.update_or_create(nombre=nombre, defaults=defaults)
                        
                        if created: summary["created"] += 1
                        else: summary["updated"] += 1

        # --- Subbodegas ---
        if "Subbodegas" in wb.sheetnames:
            ws = wb["Subbodegas"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "ID")
                    nombre = get_clean_val(row, headers, "Nombre")
                    bodega_val = get_clean_val(row, headers, "Bodega Padre")
                    parent_val = get_clean_val(row, headers, "Subbodega Padre")
                    
                    if nombre and bodega_val:
                        bodega_obj = Bodega.objects.filter(nombre=bodega_val).first()
                        if bodega_obj:
                            parent_obj = None
                            if parent_val:
                                parent_obj = Subbodega.objects.filter(nombre=parent_val, bodega=bodega_obj).first()
                            
                            defaults = {
                                'nombre': nombre, 
                                'bodega': bodega_obj, 
                                'parent': parent_obj,
                                'activo': True
                            }
                            if id_val:
                                obj, created = Subbodega.objects.update_or_create(id=id_val, defaults=defaults)
                            else:
                                # Natural key fallback: nombre + bodega
                                obj, created = Subbodega.objects.update_or_create(nombre=nombre, bodega=bodega_obj, defaults=defaults)
                            
                            if created: summary["created"] += 1
                            else: summary["updated"] += 1

        # --- Materiales ---
        if "Materiales" in wb.sheetnames:
            ws = wb["Materiales"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "ID")
                    codigo = get_clean_val(row, headers, "Código")
                    nombre = get_clean_val(row, headers, "Nombre")
                    if codigo and nombre:
                        marca_val = get_clean_val(row, headers, "Marca")
                        marca_obj = None
                        if marca_val:
                            # For Brand lookup in Material sheets, we assume it's by name
                            marca_obj = Marca.objects.filter(nombre=marca_val).first()
                        
                        defaults = {
                            'codigo': codigo,
                            'codigo_barras': get_clean_val(row, headers, "Código Barras"),
                            'referencia': get_clean_val(row, headers, "Referencia"),
                            'nombre': nombre,
                            'unidad': get_clean_val(row, headers, "Unidad") or "und",
                            'marca': marca_obj
                        }
                        if id_val:
                            obj, created = Material.objects.update_or_create(id=id_val, defaults=defaults)
                        else:
                            # Natural key fallback: codigo
                            obj, created = Material.objects.update_or_create(codigo=codigo, defaults=defaults)
                        
                        if created: summary["created"] += 1
                        else: summary["updated"] += 1

        # --- Facturas ---
        if "Facturas" in wb.sheetnames:
            ws = wb["Facturas"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "id")
                    numero = get_clean_val(row, headers, "numero")
                    if numero:
                        fecha_val = get_clean_val(row, headers, "fecha")
                        if isinstance(fecha_val, str):
                            try: fecha_val = datetime.datetime.strptime(fecha_val, '%Y-%m-%d').date()
                            except: fecha_val = timezone.now().date()
                        
                        defaults = {
                            'numero': numero,
                            'proveedor': get_clean_val(row, headers, "proveedor") or "",
                            'fecha': fecha_val if fecha_val else timezone.now().date()
                        }
                        if id_val:
                            obj, created = Factura.objects.update_or_create(id=id_val, defaults=defaults)
                        else:
                            # Natural key fallback: numero
                            obj, created = Factura.objects.update_or_create(numero=numero, defaults=defaults)
                        
                        if created: summary["created"] += 1
                        else: summary["updated"] += 1

        # --- Movimientos ---
        if "Movimientos" in wb.sheetnames:
            ws = wb["Movimientos"]
            rows = list(ws.rows)
            if len(rows) > 1:
                headers = [str(cell.value) for cell in rows[0]]
                for row_cells in rows[1:]:
                    row = [cell.value for cell in row_cells]
                    id_val = get_clean_val(row, headers, "ID")
                    tipo = get_clean_val(row, headers, "Tipo")
                    material_val = get_clean_val(row, headers, "Material")
                    cantidad = get_clean_val(row, headers, "Cantidad")
                    bodega_val = get_clean_val(row, headers, "Bodega")
                    
                    if tipo and material_val and cantidad is not None and bodega_val:
                        # Resolve material (Material column has "codigo - nombre")
                        mat_code = str(material_val).split(' - ')[0]
                        mat_obj = Material.objects.filter(codigo=mat_code).first()
                        
                        bod_obj = Bodega.objects.filter(nombre=bodega_val).first()
                        
                        if mat_obj and bod_obj:
                            # Resolve optional subbodega
                            sub_val = get_clean_val(row, headers, "Subbodega")
                            sub_obj = None
                            if sub_val:
                                sub_obj = Subbodega.objects.filter(nombre=sub_val, bodega=bod_obj).first()
                            
                            # Resolve destination for traslados
                            bod_dest_val = get_clean_val(row, headers, "Bodega Destino")
                            bod_dest_obj = None
                            if bod_dest_val:
                                bod_dest_obj = Bodega.objects.filter(nombre=bod_dest_val).first()
                            
                            sub_dest_val = get_clean_val(row, headers, "Subbodega Destino")
                            sub_dest_obj = None
                            if sub_dest_val and bod_dest_obj:
                                sub_dest_obj = Subbodega.objects.filter(nombre=sub_dest_val, bodega=bod_dest_obj).first()

                            # Resolve Marca
                            marca_val = get_clean_val(row, headers, "Marca")
                            marca_obj = None
                            if marca_val:
                                marca_obj = Marca.objects.filter(nombre=marca_val).first()

                            # Date parsing
                            fecha_val = get_clean_val(row, headers, "Fecha")
                            if isinstance(fecha_val, str):
                                try: fecha_val = timezone.datetime.fromisoformat(fecha_val)
                                except: fecha_val = timezone.now()
                            elif not fecha_val:
                                fecha_val = timezone.now()

                            defaults = {
                                'tipo': tipo,
                                'material': mat_obj,
                                'cantidad': int(float(cantidad)),
                                'bodega': bod_obj,
                                'subbodega': sub_obj,
                                'bodega_destino': bod_dest_obj,
                                'subbodega_destino': sub_dest_obj,
                                'marca': marca_obj,
                                'fecha': fecha_val,
                                'factura_manual': get_clean_val(row, headers, "Factura Manual"),
                                'observaciones': get_clean_val(row, headers, "Observaciones"),
                                'usuario': user
                            }
                            
                            if id_val:
                                obj, created = Movimiento.objects.update_or_create(id=id_val, defaults=defaults)
                            else:
                                # New movement row (creation)
                                obj = Movimiento.objects.create(**defaults)
                                created = True
                            
                            if created: summary["created"] += 1
                            else: summary["updated"] += 1
                        else:
                            summary["errors"].append(f"No se encontró Material '{material_val}' o Bodega '{bodega_val}' para una fila.")

    return summary
